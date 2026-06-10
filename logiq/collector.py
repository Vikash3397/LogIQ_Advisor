"""Agent 1 - Log Collector: deterministic multi-format log parsing.

Supports ``.txt``/``.log`` (line-oriented text with multi-line stack traces),
``.csv`` and ``.xlsx`` (tabular logs). Detects error/warning/exception/critical
entries, captures surrounding context, and emits :class:`ErrorRecord` objects.

The agent (LLM) can enrich or re-rank these results, but the heavy lifting of
scanning large files is done here so it is fast and reproducible.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable, List, Optional, Sequence

from .schemas import ErrorRecord

# Filename extensions handled by the text vs tabular parsers.
TEXT_EXTS = {".txt", ".log"}
TABULAR_EXTS = {".csv", ".xlsx"}
SUPPORTED_EXTS = TEXT_EXTS | TABULAR_EXTS

# Number of surrounding lines kept as context for each detected error.
CONTEXT_BEFORE = 3
CONTEXT_AFTER = 3

# Default levels collected when no explicit filter is provided.
DEFAULT_COLLECT_LEVELS = ("ERROR", "EXCEPTION", "CRITICAL", "FATAL")

# Numeric LOG_LEVEL_ID values used in DM / Financial Accounting Manager job logs.
_DM_LOG_LEVEL_MAP = {
    "1": "CRITICAL",
    "2": "ERROR",
    "4": "WARNING",
}

# Level keywords mapped to canonical levels (see schemas.ERROR_LEVELS).
_LEVEL_PATTERNS = [
    ("CRITICAL", re.compile(r"\b(CRITICAL|FATAL)\b", re.IGNORECASE)),
    ("EXCEPTION", re.compile(r"\b(EXCEPTION|TRACEBACK)\b", re.IGNORECASE)),
    ("ERROR", re.compile(r"\b(ERROR|ERR|SEVERE|FAILED|FAILURE)\b", re.IGNORECASE)),
    ("WARNING", re.compile(r"\b(WARNING|WARN)\b", re.IGNORECASE)),
]

# ISO-ish or common log timestamps, e.g. 2026-06-09 20:15:01,123 or 2026-06-09T20:15:01Z
_TIMESTAMP_RE = re.compile(
    r"(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}(?:[.,]\d{1,6})?(?:Z|[+-]\d{2}:?\d{2})?)"
)

# Logger / module name, e.g. "[com.acme.OrderService]" or "com.acme.module -" tokens.
_MODULE_BRACKET_RE = re.compile(r"\[([\w.$-]+(?:\.[\w.$-]+)*)\]")
_MODULE_DOTTED_RE = re.compile(r"\b([a-zA-Z_][\w]*(?:\.[a-zA-Z_][\w]*){2,})\b")

# A line that looks like part of a stack trace / traceback continuation.
_STACK_LINE_RE = re.compile(
    r"^\s*(at\s+[\w$.]+\(|File\s+\"|Traceback|Caused by:|\.{3}\s|"
    r"[\w.$]+(?:Error|Exception)(?::|$)|\tat\s)"
)


def detect_level(line: str) -> Optional[str]:
    """Return the canonical level for a line, or ``None`` if not an alert line."""
    for level, pattern in _LEVEL_PATTERNS:
        if pattern.search(line):
            return level
    return None


def detect_level_from_code(level_text: str) -> Optional[str]:
    """Map numeric or textual level codes to canonical levels."""
    text = level_text.strip()
    if not text:
        return None
    mapped = _DM_LOG_LEVEL_MAP.get(text)
    if mapped:
        return mapped
    return detect_level(text)


def extract_timestamp(line: str) -> Optional[str]:
    match = _TIMESTAMP_RE.search(line)
    return match.group(1) if match else None


def extract_module(line: str) -> Optional[str]:
    bracket = _MODULE_BRACKET_RE.search(line)
    if bracket:
        candidate = bracket.group(1)
        if not _TIMESTAMP_RE.search(candidate):
            return candidate
    dotted = _MODULE_DOTTED_RE.search(line)
    if dotted:
        return dotted.group(1)
    return None


def _is_stack_continuation(line: str) -> bool:
    return bool(_STACK_LINE_RE.search(line)) and detect_level(line) is None


def parse_text_lines(
    lines: Sequence[str],
    source_file: str,
    id_prefix: str,
    levels: Optional[Iterable[str]] = None,
) -> List[ErrorRecord]:
    """Parse line-oriented text logs into error records.

    Consecutive stack-trace lines following an alert line are folded into that
    record's ``stack_trace``.
    """
    wanted = {lv.upper() for lv in levels} if levels else None
    records: List[ErrorRecord] = []
    n = len(lines)
    i = 0
    seq = 0
    while i < n:
        raw = lines[i].rstrip("\n")
        level = detect_level(raw)
        if level is None or (wanted and level not in wanted):
            i += 1
            continue

        # Collect following stack-trace continuation lines.
        stack: List[str] = []
        j = i + 1
        while j < n and _is_stack_continuation(lines[j].rstrip("\n")):
            stack.append(lines[j].rstrip("\n"))
            j += 1

        before = [l.rstrip("\n") for l in lines[max(0, i - CONTEXT_BEFORE):i]]
        after_start = j
        after = [l.rstrip("\n") for l in lines[after_start:after_start + CONTEXT_AFTER]]

        seq += 1
        records.append(
            ErrorRecord(
                id=f"{id_prefix}-{seq:04d}",
                level=level,
                message=raw.strip(),
                source_file=source_file,
                line_no=i + 1,
                timestamp=extract_timestamp(raw),
                module=extract_module(raw),
                stack_trace="\n".join(stack) if stack else None,
                context_lines=before + after,
            )
        )
        i = j
    return records


def _norm_header(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).lower())


def _pick_column(headers: Sequence[str], candidates: Sequence[str]) -> Optional[int]:
    normalized = [_norm_header(h) for h in headers]
    for cand in candidates:
        if cand in normalized:
            return normalized.index(cand)
    return None


def parse_tabular_rows(
    headers: Sequence[str],
    rows: Iterable[Sequence],
    source_file: str,
    id_prefix: str,
    levels: Optional[Iterable[str]] = None,
) -> List[ErrorRecord]:
    """Parse tabular logs (CSV/XLSX) into error records.

    Recognizes common column names for level, message, timestamp, module, and
    stack trace. Rows whose level is not an alert level are skipped.
    """
    wanted = {lv.upper() for lv in levels} if levels else None
    level_idx = _pick_column(headers, ["level", "severity", "loglevel", "loglevelid", "type"])
    msg_idx = _pick_column(headers, ["message", "msg", "description", "error", "text"])
    err_msg_idx = _pick_column(headers, ["errormessage", "error_message"])
    ts_idx = _pick_column(
        headers, ["timestamp", "time", "datetime", "date", "ts", "logtimestamp"]
    )
    mod_idx = _pick_column(
        headers, ["module", "service", "logger", "component", "source", "jobprocess", "jobname"]
    )
    stack_idx = _pick_column(headers, ["stacktrace", "stack", "trace", "exception", "traceback"])

    records: List[ErrorRecord] = []
    seq = 0
    for row_no, row in enumerate(rows, start=2):  # row 1 is the header
        cells = ["" if c is None else str(c) for c in row]
        if not any(c.strip() for c in cells):
            continue

        def cell(idx: Optional[int]) -> str:
            return cells[idx].strip() if idx is not None and idx < len(cells) else ""

        level_text = cell(level_idx)
        joined = " ".join(cells)
        error_message = cell(err_msg_idx)
        level = (
            detect_level_from_code(level_text)
            or detect_level(level_text)
            or detect_level(joined)
            or ("ERROR" if error_message else None)
        )
        if level is None or (wanted and level not in wanted):
            continue

        message = error_message or cell(msg_idx) or joined.strip()
        seq += 1
        records.append(
            ErrorRecord(
                id=f"{id_prefix}-{seq:04d}",
                level=level,
                message=message,
                source_file=source_file,
                line_no=row_no,
                timestamp=cell(ts_idx) or extract_timestamp(joined) or None,
                module=cell(mod_idx) or extract_module(joined) or None,
                stack_trace=cell(stack_idx) or None,
                context_lines=[],
            )
        )
    return records


def parse_csv_file(path: Path, id_prefix: str, levels=None) -> List[ErrorRecord]:
    with path.open("r", encoding="utf-8", newline="", errors="replace") as fh:
        reader = csv.reader(fh)
        all_rows = list(reader)
    if not all_rows:
        return []
    headers = all_rows[0]
    return parse_tabular_rows(headers, all_rows[1:], path.name, id_prefix, levels)


def parse_xlsx_file(path: Path, id_prefix: str, levels=None) -> List[ErrorRecord]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError(
            "Parsing .xlsx requires 'openpyxl'. Install it with "
            "'pip install -r requirements.txt'."
        ) from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    records: List[ErrorRecord] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [("" if h is None else str(h)) for h in rows[0]]
        sheet_prefix = f"{id_prefix}-{_norm_header(sheet.title) or 'sheet'}"
        records.extend(
            parse_tabular_rows(headers, rows[1:], f"{path.name}::{sheet.title}", sheet_prefix, levels)
        )
    wb.close()
    return records


def _looks_like_csv_file(path: Path) -> bool:
    """Return True when a text file is CSV-shaped (e.g. DM job export with .txt ext)."""
    with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
        sample = fh.read(4096)
    if not sample.strip():
        return False
    try:
        rows = list(csv.reader([sample.splitlines()[0]]))
    except csv.Error:
        return False
    if not rows or not rows[0]:
        return False
    normalized = {_norm_header(h) for h in rows[0]}
    tabular_markers = {
        "loglevelid",
        "errormessage",
        "logtimestamp",
        "jobprocess",
        "jobname",
        "level",
        "severity",
        "message",
    }
    return len(normalized & tabular_markers) >= 2


def parse_text_file(path: Path, id_prefix: str, levels=None) -> List[ErrorRecord]:
    if _looks_like_csv_file(path):
        return parse_csv_file(path, id_prefix, levels)
    with path.open("r", encoding="utf-8", errors="replace") as fh:
        lines = fh.readlines()
    return parse_text_lines(lines, path.name, id_prefix, levels)


def parse_file(path: Path, id_prefix: Optional[str] = None, levels=None) -> List[ErrorRecord]:
    """Parse a single log file by extension."""
    ext = path.suffix.lower()
    prefix = id_prefix or _norm_header(path.stem) or "err"
    if ext in TEXT_EXTS:
        return parse_text_file(path, prefix, levels)
    if ext == ".csv":
        return parse_csv_file(path, prefix, levels)
    if ext == ".xlsx":
        return parse_xlsx_file(path, prefix, levels)
    raise ValueError(f"Unsupported file type '{ext}' for {path.name}")


def _iter_input_files(target: Path) -> List[Path]:
    if target.is_file():
        return [target]
    if target.is_dir():
        return sorted(
            p for p in target.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_EXTS
        )
    raise FileNotFoundError(f"Input path does not exist: {target}")


def parse_path(input_path: str, levels=None) -> List[ErrorRecord]:
    """Parse a file or directory of logs into a flat list of error records.

    When parsing a directory, the per-file id prefix is derived from the file
    name so ids remain unique and traceable across the run.

    By default only ERROR, EXCEPTION, CRITICAL, and FATAL entries are kept;
    WARNING and informational messages are excluded unless ``levels`` is overridden.
    """
    effective_levels = DEFAULT_COLLECT_LEVELS if levels is None else levels
    target = Path(input_path)
    files = _iter_input_files(target)
    records: List[ErrorRecord] = []
    used_prefixes: dict = {}
    for f in files:
        base = _norm_header(f.stem) or "err"
        used_prefixes[base] = used_prefixes.get(base, 0) + 1
        prefix = base if used_prefixes[base] == 1 else f"{base}{used_prefixes[base]}"
        records.extend(parse_file(f, id_prefix=prefix, levels=effective_levels))
    return records
